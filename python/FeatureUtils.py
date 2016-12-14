# -*- coding: utf-8 -*-
"""
Created on Sun Oct 04 21:37:44 2015

@author: Hongjian

Package name: featureUtil

various functions for generating feature matrices/vectors


"""

from Crime import Tract
import numpy as np
import csv
from openpyxl import load_workbook
import heapq
import pickle


import os
here = os.path.dirname(os.path.abspath(__file__))



def generate_corina_features(region='ca', leaveOut=-1):
    """
    Generate the features recommended by Corina.
    
    parameter region taks 'ca' or 'tract'
    
    Return values are field description, value array
    """
    if region == 'ca':
        f = open(here + '/../data/Chicago_demographics.csv', 'r')
        c = csv.reader(f)
        header = c.next()
        fields = ['totpop00_sum', 'popden00', 'pprpovW00', 'Dis46pf0', 'Stb26pf0', 'Divers5f00', 
                'pnhblWc0', 'phispWc0']
        fields_dsp = ['total population', 'population density', 'poverty index', 'disadvantage index', 'residential stability',
                'ethnic diversity', 'pct black', 'pct hispanic']
        hidx = []
        for fd in fields:
            hidx.append( header.index(fd) )
        
        C = np.zeros( (77,len(hidx)) )
        for i, row in enumerate(c):
            for j, k in enumerate( hidx ):
                C[i][j] = float(row[k])
                
        if leaveOut > 0:
            C = np.delete(C, leaveOut-1, 0)
    
        return  fields_dsp, C
    elif region == 'tract':
        
        from pandas import read_stata
    
        r = read_stata('../data/SE2000_AG20140401_MSAcmsaID.dta')
        cnt = 0
        header = ['pop00', 'ppov00', 'disadv00', 'pdensmi00', 'hetero00', 'phisp00', 'pnhblk00']
        
        fields_dsp = ['total population', 'poverty index', 'disadvantage index', 'population density',
                'ethnic diversity', 'pct hispanic', 'pct black']
                
        ST = {}
        for row in r.iterrows():
            tract = row[1]
            if tract['statetrim'] == '17' and tract['countrim'] == '031':
                cnt += 1
                tl = []
                tid = long('17031' + tract['tracttrim'])
                for h in header:
                    tl.append(tract[h])
                ST[tid] = tl
        return fields_dsp, ST



def generate_geographical_SpatialLag():
    """
    Generate the spatial lag from the geographically adjacent regions.
    """
    ts = Tract.createAllTractObjects()
    ordkey = sorted(ts)
    centers = [ ts[k].polygon.centroid for k in ordkey ]
    
    W = np.zeros( (len(centers), len(centers)) )
    for i, src in enumerate(centers):
        for j, dst in enumerate(centers):
            if src != dst:
                W[i][j] = 1 / src.distance(dst)
    return W, ordkey
        
        

def generate_geographical_SpatialLag_ca(knearest=True, leaveOut=-1):
    """
    Generate the distance matrix for CA pairs.
    
    If knearest is true, then select the 6-nearest neighboring CAs.
    Else, return the distance to all other CAs.

    leaveOut will select the CA and remove it. take value from 1 to 77
    """
    
    
    cas = Tract.createAllCAObjects()
    centers = []
    iset = range(1, 78)
    if leaveOut > 0:
        iset.remove(leaveOut)
    for i in iset:
        centers.append(cas[i].polygon.centroid)
    
    W = np.zeros( (len(iset),len(iset)) )
    for i, src in enumerate(centers):
        for j, dst in enumerate(centers):
            if src != dst:
                W[i][j] = 1 / src.distance(dst)
                
        # find n-largest (n=6)
        if knearest == True:
            threshold = heapq.nlargest(6, W[i,:])[-1]
            for j in range(len(W[i,:])):
                W[i][j] = 0 if W[i][j] < threshold else W[i][j]
    return W
    
        

def generate_GWR_weight(h = 1):
    """
    Generate the GWR weighting matrix with exponential function.
    """
    cas = Tract.createAllCAObjects()
    centers = []
    for i in range(1, 78):
        centers.append(cas[i].polygon.centroid)
    
    gamma = np.ones((len(centers), len(centers)))
    for i, src in enumerate(centers):
        for j, dst in enumerate(centers):
            if i != j:
                gamma[i][j] = np.exp(-0.5 * src.distance(dst)**2 / h**2)
    return gamma
    
    
def generate_geo_graph_embedding_src():
    flow = generate_geographical_SpatialLag_ca()
    row, column = flow.shape
    with open("multi-view-learning/geo.od", 'w') as fout:
        for i in range(row):
            for j in range(column):
                if flow[i,j] > 0:
                    fout.write('{0} {1} {2}\n'.format(i, j, flow[i,j]))


                    
def generate_transition_SocialLag(year = 2010, lehd_type=0, region='ca', leaveOut=-1, normalization='source'):
    """
    Generate the spatial lag matrix from the transition flow connected CAs.
    
    0 - #total jobs
    1 - #jobs age under 29,
    2 - #jobs age from 30 to 54, 
    3 - #jobs above 55, 
    4 - #jobs earning under $1250/month, 
    5 - #jobs earnings from $1251 to $3333/month, 
    6 - #jobs above $3333/month,
    7 - #jobs in goods producing, 
    8 - #jobs in trade transportation, 
    9 - #jobs in other services
    """
    
    if region == 'ca':
        ts = Tract.createAllCAObjects()
        fn = here + '/../data/chicago_ca_od_{0}.csv'.format(year)
    elif region == 'tract':
        ts = Tract.createAllTractObjects()
        fn = here + '/../data/chicago_od_tract_{0}.csv'.format(year)
    ordkey = sorted(ts.keys())
    
    
    listIdx = {}
    fin = open(fn)
    for line in fin:
        ls = line.split(",")
        srcid = int(ls[0])
        dstid = int(ls[1])
        val = int(ls[2 + lehd_type])
        if srcid in listIdx:
            listIdx[srcid][dstid] = val
        else:
            listIdx[srcid] = {}
            listIdx[srcid][dstid] = val                            
    fin.close()

    if leaveOut > 0:
        ordkey.remove(leaveOut)
    
    W = np.zeros( (len(ts),len(ts)) )
    for srcid in ordkey:
        if srcid in listIdx:
            sdict = listIdx[srcid]
            if leaveOut in sdict:
                del sdict[leaveOut]
            for dstid, val in sdict.items():
                W[ordkey.index(srcid)][ordkey.index(dstid)] = val
        else:
            W[ordkey.index(srcid)] = np.zeros( (1,len(ts)) )
            
    
    
    # update diagonal as 0
#    if normalization != 'none':
#        for i in range(len(W)):
#            W[i,i] = 0
    # first make all self-factor 0
    assert W.dtype == "float64"
        
    # normalization section
    if normalization == 'source':
        # source mean the residence
        W = np.transpose(W)    
        sW = np.sum(W, axis=1, keepdims=True)
        W = W / sW
        assert abs( np.sum(W[1,]) - 1 ) < 0.0000000001 and W.dtype == "float64"
    elif normalization == 'destination': # 
        # destination mean workplace
        sW = np.sum(W, axis=1)
        sW = sW.reshape((len(sW),1))
        W = W / sW
    elif normalization == 'pair':
        sW = W + np.transpose(W)
        sW = np.sum(sW)
        W = W / sW
    
    # by default, the output is the workplace-to-residence count matrix
    return W


import pandas as pd

def retrieve_health_data():
    """
    get health data
    """
    h = pd.read_stata("../data/ChiCas77_PubHealthScale_ForNSFproject.dta")
    return h['phlth12vc2alpha'].values, h['phlth10novcAlpha'].values



def retrieve_crime_count(year, col=['total'], region='ca'):
    """
    Retrieve the crime count in a vector
    Input:
        year - the year to retrieve
        col  - the type of crime
        region - ca or tract
        
    Output:
        if region == 'ca':  Y is a column vector of size (77,1)
    """
    if region == 'ca':
        Y =np.zeros( (77,1) )
        with open(here + '/../data/chicago-crime-ca-level-{0}.csv'.format(year)) as fin:
            header = fin.readline().strip().split(",")
            crime_idx = []
            for c in col:
                if c in header:
                    i = header.index(c)
                    crime_idx.append(i)
            for line in fin:
                ls = line.split(",")
                idx = int(ls[0])
                val = 0
                for i in crime_idx:
                    val += int(ls[i])
                Y[idx-1] = val
    
        return Y
        
    elif region == 'tract':
        Y = {}
        with open(here + '/../data/chicago-crime-tract-level-{0}.csv'.format(year)) as fin:
            header = fin.readline().strip().split(",")
            crime_idx = []
            for c in col:
                i = header.index(c)
                crime_idx.append(i)
            for line in fin:
                ls = line.split(",")
                tid = int(ls[0])
                val = 0
                for i in crime_idx:
                    val += int(ls[i])
                Y[tid] = val
        return Y




def retrieve_income_features():
    """
    read the xlsx file: ../data/chicago-ca-income.xlsx
    
    Three kinds of features we can generate: 
    1. population count in each category
    2. probability distribution over all categories (normalize by population)
    3. Grouped mean, variance    
    """
    wb = load_workbook(here + "/../data/chicago-ca-income.xlsx")
    ws = wb.active
    header = ws['l3':'aa3']
    header = [c.value for c in tuple(header)[0]]
    
#    bins = [5000, 12500, 17500, 22500, 27500, 32500, 37500, 42500, 47500, 55000, 67500,
#            87500, 112500, 137500, 175000, 300000]
    bins = range(1,17)
    l = len(header)
    I = np.zeros((77,l))
    stats_header = ['income mean', 'std var']
    stats = np.zeros((77,2))    # mean, variance
    total = np.zeros( (77,1) )
    for idx, row in enumerate(ws.iter_rows('k4:aa80')):
        for j, c in enumerate(row):
            if j == 0:
                total[idx] =  float(c.value)
            else:
                I[idx][j-1] = c.value # / total
        stats[idx][0] = np.dot(bins, I[idx][:]) / total[idx]
        stats[idx][1] = np.sqrt( np.dot(I[idx][:], (bins - stats[idx][0])**2) / total[idx] )
#    return header, I
    return stats_header, stats, ['total'], total





def retrieve_education_features():
    """
    read the xlsx file: ../data/chicago-ca-education.xlsx
    """
    wb = load_workbook(here + "/../data/chicago-ca-education.xlsx")
    ws = wb.active
    header = ws['k3':'n3']
    header = [c.value for c in tuple(header)[0]]
    
    bins = range(1,5)
    l = len(header)
    E = np.zeros((77,l))
    stats_header = ['education level', 'std var']
    stats = np.zeros((77,2))
    for i, row in enumerate(ws.iter_rows('j4:n80')):
        total = 0
        for j, c in enumerate(row):
            if j == 0:
                total = float(c.value)
            else:
                E[i][j-1] = c.value # / total
        stats[i][0] = np.dot(E[i][:], bins) / total
        stats[i][1] = np.sqrt( np.dot(E[i][:], (bins - stats[i][0])**2) / total)
    return stats_header, stats
                    
        
    
    
    
    
def retrieve_race_features():
    """
    read the xlsx file: ../data/chicago-ca-race.xlsx
    """
    wb = load_workbook(here + "/../data/chicago-ca-race.xlsx")
    ws = wb.active
    header = ws['j2':'p2']
    header = [c.value for c in tuple(header)[0]]
    l = len(header)
    R = np.zeros((77,l))
    
    bins = range(1,8)
    
    stats_header = ['race level', 'std var']
    stats = np.zeros((77,2))
    for i, row in enumerate(ws.iter_rows('j4:p80')):
        total = 0
        for c in row:
            total += float(c.value)
        for j, c in enumerate(row):
            R[i][j] = c.value # / total
        
        stats[i][0] = np.dot(R[i][:], bins) / total
        stats[i][1] = np.sqrt( np.dot(R[i][:], (bins - stats[i][0])**2) / total)
    return stats_header, stats
#    return header, R
    
    
    
    
def generateDotFile(s, threshold=400, fileName='taxiflow'):
    """
    generate dot file
    
    The dot file is used by graphviz to visualize graph
    """
    nodes = set()
    with open('{0}.dot'.format(fileName), 'w') as fout:
        fout.write('digraph taxiChicago {\n')
        for i, row in enumerate(s):
            for j, ele in enumerate(row):
                if ele > threshold:
                    if i not in nodes:
                        fout.write('{0} [label="{1}"];\n'.format(i, i+1))
                        nodes.add(i)
                    if j not in nodes:
                        fout.write('{0} [label="{1}"];\n'.format(j, j+1))
                        nodes.add(j)
                    fout.write('{0} -> {1} [label="{2:.3f}"];\n'.format(i,j, ele) )
        fout.write('}\n')
        
    import subprocess
    subprocess.call(['dot ', '-Tpdf', '-o{0}.pdf'.format(fileName),  '{0}.dot'.format(fileName)])
    


import unittest

class TestFeatureUtils(unittest.TestCase):
    
    def test_generate_GWR_weight(self):
        gamma = generate_GWR_weight(0.5)
        for i in range(20):
            np.testing.assert_almost_equal(gamma[i,i], 1.)
        assert np.amax(gamma) <= 1
        print np.amin(gamma)
        
        
def generate_binary_crime_label():
    y = retrieve_crime_count(2013)
    threshold = np.median(y)
    label = [1 if ele >= threshold else 0 for ele in y]
    F = generate_corina_features()
    from sklearn import svm, tree
    from sklearn.model_selection import cross_val_score
    clf1 = svm.SVC()
    scores1 = cross_val_score(clf1, F[1], label, cv=10)
    print scores1.mean(), scores1
    clf2 = tree.DecisionTreeClassifier()
    scores2 = cross_val_score(clf2, F[1], label, cv=10)
    print scores2.mean(), scores2
    pickle.dump(label, open("crime-label", 'w'))
    return y, label, F[1]
    
    
def generate_binary_demo_label():
    D = generate_corina_features()
    demolabel = {}
    for i, d in enumerate(D[0]):
        F = D[1][:,i]
        thrsd = np.median(F)
        label= [1 if ele >= thrsd else 0 for ele in F]
        demolabel[d] = label
    pickle.dump(demolabel, open("demo-label", "w"))
    return demolabel, D
    
    
    

def generate_lehd_label():
    f = generate_transition_SocialLag(2010, 0, 'ca', -1, 'None')
    r = []
    for i in range(77):
        home = np.sum(f[:,i])
        work = np.sum(f[i,:])
        r.append(home/work)
    avg, std = np.mean(r), np.std(r)
    label = []
    cnt = [0,0,0]
    for i in r:
        if i >= avg + std / 2:
            label.append(1)
            cnt[0] += 1
        elif i <= avg - std / 2:
            label.append(-1)
            cnt[1] += 1
        else:
            label.append(0)
            cnt[2] += 1
    for idx, i in enumerate(r):
        print idx+1, i, '\t', label[idx]
    print cnt
    pickle.dump(label, open("lehd-label", "w"))
    return f, label

    
    

if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1: 
        if sys.argv[1] == 'test':
            unittest.main()
        elif sys.argv[1] == 'binarylabel':
            y, l, f = generate_binary_crime_label()
            l, D = generate_binary_demo_label()
        elif sys.argv[1] == 'LEHDlabel':
            l = generate_lehd_label()
    else:
        generate_geo_graph_embedding_src()
    
#    from taxiFlow import getTaxiFlow
#    s = getTaxiFlow(usePercentage=False)
#    generateDotFile(s, 5000)
    
    
#    for year in range(2002, 2014):
#        t = generate_transition_SocialLag(year=year, lehd_type=0, region='ca', 
#                                      leaveOut=-1, normalization='none')
#        np.savetxt("{0}-social-row-matrix.csv".format(year), t, delimiter=",")
#    generateDotFile(t, 0.08, 'sociallag')
    
#    h = retrieve_health_data()
