# -*- coding: utf-8 -*-
"""
Created on Wed Sep  2 14:24:07 2015

@author: hxw186


extract CA level features from xlsx file.
"""

from openpyxl import *

def extractCAlevelFeature(fin = "../data/Chicago_race_place of origin _census_tract_CA_2010.xlsx", fout = '../data/chicago-ca-race.xlsx'):
    wb = load_workbook(fin)
   
    wo = Workbook()
    wos = wo.active
    
    ws = wb.active
    
    for row in ws.iter_rows():
        if row[0].row in [1, 2, 3]:
            wos.append( [c.value for c in row]  )
        elif row[1].value == '070':
            wos.append( [c.value for c in row]  )
            
            
    wo.save(filename=fout)
    return wo



if __name__ == '__main__':
    
    extractCAlevelFeature("../data/Education by Race by Census Tract and Community Area.xlsx", "../data/chicago-ca-education.xlsx")
    w = extractCAlevelFeature("../data/Household Income by Race and Census Tract and Community Area.xlsx", "../data/chicago-ca-income.xlsx")
    extractCAlevelFeature()