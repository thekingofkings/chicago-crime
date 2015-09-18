"""
Run the Negative Binomial Regression model

The regresion model use various features to predict the crime count in each
unit of study area (i.e. tract or Community Area)

Author: Hongjian
date:8/20/2015
"""

from Crime import Tract
from sets import Set
from shapely.geometry import MultiLineString
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
from scipy.stats import nbinom
from scipy.special import gammaln
from statsmodels.base.model import GenericLikelihoodModel
from openpyxl import *
import pandas as pd
from sklearn.preprocessing import scale
import statsmodels.api as sm
from sklearn import cross_validation
import csv
import subprocess
import os.path




"""
Part One
Generate vairous features
"""


def generate_corina_features():
    """
    Generate the features recommended by Corina
    """
    f = open('../data/Chicago_demographics.csv', 'r')
    c = csv.reader(f)
    header = c.next()
    fields = ['totpop00_sum', 'popden00', 'pprpovW00', 'Dis46pf0', 'Stb26pf0', 'Divers5f00', 
            'pnhblWc0', 'phispWc0']
    fields_dsp = ['total population', 'population density', 'poverty index', 'disadvantage index', 'residential stability',
            'ethnic diversity', 'pct black', 'hispanic']
    hidx = []
    for fd in fields:
        hidx.append( header.index(fd) )
    
    C = np.zeros( (77,len(hidx)) )
    for i, row in enumerate(c):
        for j, k in enumerate( hidx ):
            C[i][j] = float(row[k])

    return  fields_dsp, C



def generate_geographical_SpatialLag(foutName):
    """
    Generate the spatial lag from the geographically adjacent CAs.
    """
    fout = open(foutName, 'w')
    cnt = 0
    ts = Tract.createAllTractObjects()
    idset = Set(ts.keys())
    for i in ts.keys():
        idset.remove(i)
        for j in idset:
            if type(ts[i].polygon.intersection(ts[j].polygon)) is MultiLineString:
                fout.write('{0},{1}\n'.format(i,j))
                cnt += 1
    fout.close()
    return cnt
        
        

def generate_geographical_SpatialLag_ca():
    
    
    cas = Tract.createAllCAObjects()
    centers = []
    for i in range(1,78):
        centers.append(cas[i].polygon.centroid)
    
    W = np.zeros( (77,77) )
    for i, src in enumerate(centers):
        for j, dst in enumerate(centers):
            if src != dst:
                W[i][j] = src.distance(dst)
    return W
    
        
        


def generate_transition_SocialLag(year = 2010, lehd_type=0):
    """
    Generate the spatial lag from the transition flow connected CAs.
    
    0 - #jobs age under 29, 
    1 - #jobs age from 30 to 54, 
    2 - #jobs above 55, 
    3 - #jobs earning under $1250/month, 
    4 - #jobs earnings from $1251 to $3333/month, 
    5 - #jobs above $3333/month,
    6 - #jobs in goods producing, 
    7 - #jobs in trade transportation, 
    8 - #jobs in other services
    """
    listIdx = {}
    fin = open('../data/chicago_ca_od_{0}.csv'.format(year))
    for line in fin:
        ls = line.split(",")
        srcid = int(ls[0])
        dstid = int(ls[1])
        val = int(ls[2 + lehd_type])
        if srcid in listIdx:
            listIdx[srcid][dstid] = val
        else:
            listIdx[srcid] = {}
            listIdx[srcid][dstid] = val                            
    fin.close()

    W = np.zeros( (77,77) )
    for srcid, sdict in listIdx.items():
        total = (float) (sum( sdict.values() ))
        for dstid, val in sdict.items():
            if srcid != dstid:
                if total == 0:
                    W[srcid-1][dstid-1] = 0
                else:
                    W[srcid-1][dstid-1] = val / total

    return W




def retrieve_crime_count(year, col=-1):
    """
    Retrieve the crime count in a vector
    Input:
        year - the year to retrieve
        col  - the type of crime
    """
    Y =np.zeros( (77,1) )
    with open('../data/chicago-crime-ca-level-{0}.csv'.format(year)) as fin:
        for line in fin:
            ls = line.split(",")
            idx = int(ls[0])
            val = int(ls[col])
            Y[idx-1] = val

    return Y




def retrieve_income_features():
    """
    read the xlsx file: ../data/chicago-ca-income.xlsx
    
    Three kinds of features we can generate: 
    1. population count in each category
    2. probability distribution over all categories (normalize by population)
    3. Grouped mean, variance    
    """
    wb = load_workbook("../data/chicago-ca-income.xlsx")
    ws = wb.active
    header = ws['l3':'aa3']
    header = [c.value for c in tuple(header)[0]]
    
#    bins = [5000, 12500, 17500, 22500, 27500, 32500, 37500, 42500, 47500, 55000, 67500,
#            87500, 112500, 137500, 175000, 300000]
    bins = range(1,17)
    l = len(header)
    I = np.zeros((77,l))
    stats_header = ['income mean', 'std var']
    stats = np.zeros((77,2))    # mean, variance
    total = np.zeros( (77,1) )
    for idx, row in enumerate(ws.iter_rows('k4:aa80')):
        bin_vals = []
        for j, c in enumerate(row):
            if j == 0:
                total[idx] =  float(c.value)
            else:
                I[idx][j-1] = c.value # / total
        stats[idx][0] = np.dot(bins, I[idx][:]) / total[idx]
        stats[idx][1] = np.sqrt( np.dot(I[idx][:], (bins - stats[idx][0])**2) / total[idx] )
#    return header, I
    return stats_header, stats, ['total'], total





def retrieve_education_features():
    """
    read the xlsx file: ../data/chicago-ca-education.xlsx
    """
    wb = load_workbook("../data/chicago-ca-education.xlsx")
    ws = wb.active
    header = ws['k3':'n3']
    header = [c.value for c in tuple(header)[0]]
    
    bins = range(1,5)
    l = len(header)
    E = np.zeros((77,l))
    stats_header = ['education level', 'std var']
    stats = np.zeros((77,2))
    for i, row in enumerate(ws.iter_rows('j4:n80')):
        total = 0
        for j, c in enumerate(row):
            if j == 0:
                total = float(c.value)
            else:
                E[i][j-1] = c.value # / total
        stats[i][0] = np.dot(E[i][:], bins) / total
        stats[i][1] = np.sqrt( np.dot(E[i][:], (bins - stats[i][0])**2) / total)
    return stats_header, stats
                    
        
    
    
    
    
def retrieve_race_features():
    """
    read the xlsx file: ../data/chicago-ca-race.xlsx
    """
    wb = load_workbook("../data/chicago-ca-race.xlsx")
    ws = wb.active
    header = ws['j2':'p2']
    header = [c.value for c in tuple(header)[0]]
    l = len(header)
    R = np.zeros((77,l))
    
    bins = range(1,8)
    
    stats_header = ['race level', 'std var']
    stats = np.zeros((77,2))
    for i, row in enumerate(ws.iter_rows('j4:p80')):
        total = 0
        for c in row:
            total += float(c.value)
        for j, c in enumerate(row):
            R[i][j] = c.value # / total
        
        stats[i][0] = np.dot(R[i][:], bins) / total
        stats[i][1] = np.sqrt( np.dot(R[i][:], (bins - stats[i][0])**2) / total)
#    return stats_header, stats
    return header, R
    
    
    


"""
Part Two
Regression models
"""

    

def linearRegression(features, Y):
    """
    learn the linear regression model from features to Y
    output the regression analysis parameters
    plot scatter plot
    """
    mod = sm.OLS(Y, features )
    res = mod.fit()
    return res


    
    

class NegBin(GenericLikelihoodModel):
    """
    negative binomial regression
    """
    
    def __init__(self, endog, exog, **kwds):
        super(NegBin, self).__init__(endog, exog, **kwds)
        
        
    def nloglikeobs(self, params):
        alpha = params[-1]
        beta = params[:-1]
        mu = np.exp(np.dot(self.exog, beta))
        size = 1 / alpha
        prob = size / (size+mu)
      #  ll = 0
        # for idx, y in enumerate(self.endog):
         #    ll += gammaln(y + size) - gammaln(size) - gammaln(y+1) + y * np.log(mu * alpha / (mu *alpha + 1))- size * np.log(mu * alpha + 1)
        ll = nbinom.logpmf( self.endog, size, prob)
        return -ll
        
    def fit(self, start_params=None, maxiter = 10000, maxfun=10000, **kwds):
        if start_params == None:
            start_params = np.append(np.zeros(self.exog.shape[1]), .5)
            if self.exog.mean() > 1:
                start_params[:-1] = np.ones(self.exog.shape[1]) * np.log(self.endog.mean())  / self.exog.mean()
            else:
                start_params[:-1] = np.ones(self.exog.shape[1]) * np.log(self.endog.mean())
#            print "endog mean:", self.endog.mean(), "log endog mean:", np.log(self.endog.mean())
#            print "exog mean", self.exog.mean()
        return super(NegBin, self).fit(start_params=start_params, maxiter=maxiter,
                maxfun=maxfun, **kwds)
                
    
    def predict(self, params, exog=None, *args, **kwargs):
        """
        predict the acutal endogenous count from exog
        """
        beta = params[:-1]
        return np.exp(exog.dot(beta))
  


    
def negativeBinomialRegression(features, Y):
    """
    learn the NB regression
    """
    mod = NegBin(Y, features)
    res = mod.fit(disp=False)
    if not res.mle_retvals['converged']:
        print "NBreg not converged.", res.params[0], ",", res.pvalues[0]
    return res, mod


    
def unitTest_withOnlineSource():
    import patsy
    import pandas as pd
    url = 'http://vincentarelbundock.github.com/Rdatasets/csv/COUNT/medpar.csv'
    medpar = pd.read_csv(url)
    y, X = patsy.dmatrices('los~type2+type3+hmo+white', medpar)
    res = negativeBinomialRegression(X, y)
    return y, X, medpar
    
    
def unitTest_onChicagoCrimeData():

    W = generate_transition_SocialLag(2010)
    Yhat = retrieve_crime_count(2009, -1)
    Y = retrieve_crime_count(2010, -1)
    i = retrieve_income_features()
    e = retrieve_education_features()
    r = retrieve_race_features()
    C = generate_corina_features()
        
    f1 = np.dot(W, Yhat)
    # f = np.concatenate((f1, i[1], e[1], r[1], np.ones(f1.shape)), axis=1)
    # f = pd.DataFrame(f, columns=['social lag'] + i[0] + e[0] + r[0] + ['intercept'])
#    f = scale(f)
#    f = np.concatenate((f, np.ones(f1.shape)), axis=1)
    f = np.concatenate( (C[1], np.ones(f1.shape)), axis=1 )
    np.savetxt("Y.csv", Y, delimiter=",")
    f = pd.DataFrame(f, columns=C[0] + ['intercept'])
    f.to_csv("f.csv", sep="," )
    Y = Y.reshape((77,))
    print "Y", Y.mean()
    res = negativeBinomialRegression(f, Y)
    
    
    print "f shape", f.shape
    print "Y shape", Y.shape
    linearRegression(f, Y)
    return res
     
    


def leaveOneOut_evaluation_onChicagoCrimeData(year=2010, features= ["all"], crime_idx=-1, flow_type=0, verboseoutput=False):
    """
    Generate the social lag from previous year
    use income/race/education of current year
    """
    W = generate_transition_SocialLag(year-1, lehd_type=flow_type)
    Yhat = retrieve_crime_count(year-1, crime_idx)
    Y = retrieve_crime_count(year, crime_idx)
    C = generate_corina_features()
    
    W2 = generate_geographical_SpatialLag_ca()
    
    i = retrieve_income_features()
    e = retrieve_education_features()
    r = retrieve_race_features()
    
    f1 = np.dot(W, Yhat)
    f2 = np.dot(W2, Yhat)
    # add intercept
    columnName = ['intercept']
    f = np.ones(f1.shape)

    if "all" in features:
        f = np.concatenate( (f, f1, i[1], e[1], r[1]), axis=1)
        f = pd.DataFrame(f, columns=['social lag'] + i[0] + e[0] + r[0])
    if "sociallag" in features: 
        f = np.concatenate( (f, f1), axis=1)
        columnName += ['social lag']
    if  "income" in features:
        f = np.concatenate( (f, i[1]), axis=1)
        columnName += i[0]
    if "race" in features:
        f = np.concatenate( (f, r[1]), axis=1)
        columnName += r[0]
    if "education" in features :
        f = np.concatenate( (f, e[1]), axis=1)
        columnName += e[0]
    if 'corina' in features :
        f = np.concatenate( (f, C[1]), axis=1)
        columnName += C[0]
    if 'spatiallag' in features:
        f = np.concatenate( (f, f2), axis=1)
        columnName += ['spatial lag']
    f = pd.DataFrame(f, columns = columnName)

        
    # call the Rscript to get Negative Binomial Regression results
    np.savetxt("Y.csv", Y, delimiter=",")
    f.to_csv("f.csv", sep=",", index=False)
    if verboseoutput:
        subprocess.call( ['Rscript', 'nbr_eval.R', 'verbose'] )
    else:
        nbres = subprocess.check_output( ['Rscript', 'nbr_eval.R'] )
    
    Y = Y.reshape((77,))
    loo = cross_validation.LeaveOneOut(77)
    mae = 0
    mae2 = 0
    errors1 = []
    errors2 = []
    for train_idx, test_idx in loo:
        f_train, f_test = f.loc[train_idx], f.loc[test_idx]
        Y_train, Y_test = Y[train_idx], Y[test_idx]
#        res, mod = negativeBinomialRegression(f_train, Y_train)
#        ybar = mod.predict(res.params, exog=f_test)
#        errors1.append( np.abs(Y_test - ybar.values[0])[0] )


        r2 = linearRegression(f_train, Y_train)
        y2 = r2.predict(f_test)
        errors2.append( np.abs( Y_test - y2 ) )
#        print test_idx, Y_test[0], ybar.values[0], y2[0]
        if verboseoutput:
            print test_idx, Y_test[0], y2[0]
        
#    mae = np.mean(errors1)
    mae2 = np.mean(errors2)
#    var = np.sqrt( np.var(errors1) )
    var2 = np.sqrt( np.var(errors2) )
#    mre = mae / Y.mean()
    mre2 = mae2 / Y.mean()
#    print "NegBio Regression MAE", mae, "std", var, "MRE", mre
    if verboseoutput:
        print "Linear Regression MAE", mae2, "std", var2, "MRE", mre2
    else:
        print nbres
        print mae2, var2, mre2
        return np.array([[float(e) for e in nbres.split(" ")], [mae2, var2, mre2]])
    





def permutationTest_onChicagoCrimeData(year=2010, features= ["all"], iters=1001):
    """
    Permutation test with regression model residuals
    """
    W = generate_transition_SocialLag(year-1)
    Yhat = retrieve_crime_count(year-1, -1)
    Y = retrieve_crime_count(year, -1)
    C = generate_corina_features()
    
    W2 = generate_geographical_SpatialLag_ca()
    
    i = retrieve_income_features()
    e = retrieve_education_features()
    r = retrieve_race_features()
    
    LR_coeffs = []
    if not os.path.exists('coefficients.txt'):
        for i in range(iters):
            if i == 0:
                pidx = range(len(Y))
            else:
                pidx = np.random.permutation(len(Y))
            f1 = np.dot(W, Yhat[pidx])
            f2 = np.dot(W2, Yhat[pidx])
            # add intercept
            columnName = ['intercept']
            f = np.ones(f1.shape)
        
            if "all" in features:
                f = np.concatenate( (f, f1, i[1], e[1], r[1]), axis=1)
                f = pd.DataFrame(f, columns=['social lag'] + i[0] + e[0] + r[0])
            if "sociallag" in features: 
                f = np.concatenate( (f, f1), axis=1)
                columnName += ['social lag']
            if  "income" in features:
                f = np.concatenate( (f, i[1]), axis=1)
                columnName += i[0]
            if "race" in features:
                f = np.concatenate( (f, r[1]), axis=1)
                columnName += r[0]
            if "education" in features :
                f = np.concatenate( (f, e[1]), axis=1)
                columnName += e[0]
            if 'corina' in features :
                f = np.concatenate( (f, C[1]), axis=1)
                columnName += C[0]
            if 'spatiallag' in features:
                f = np.concatenate( (f, f2), axis=1)
                columnName += ['spatial lag']
            f = pd.DataFrame(f, columns = columnName)
        
                
            # call the Rscript to get Negative Binomial Regression results
            np.savetxt("Y.csv", Y[pidx], delimiter=",")
            f.to_csv("f.csv", sep=",", index=False)
            subprocess.call( ['Rscript', 'nbr_permutation_test.R'] )
            
            
            # LR permutation test
            lrmod = linearRegression(f, Y)
            LR_coeffs.append(lrmod.params)
    else:        
        print 'file coefficients.txt exists!'
        
    NB_coeffs = np.loadtxt(fname='coefficients.txt', delimiter=',')
    LR_coeffs = np.array(LR_coeffs)
    
    for idx in range(NB_coeffs.shape[1]):
        column = NB_coeffs[:,idx]
        targ = column[0]
        cnt = 0.0
        for e in column:
            if e > targ:
                cnt += 1
                
        lr_col = LR_coeffs[:,idx]
        lr_trg = lr_col[0]
        lr_cnt = 0.0
        for e in lr_col:
            if e > lr_trg:
                lr_cnt += 1
                
        plt.figure(figsize=(8,3))
        # NB
        plt.subplot(1,2,1)
        plt.hist(column)
        plt.axvline(x = targ, linewidth=4, color='r')
        plt.title("NB {0} coeff pvalue {1:.4f}".format(columnName[idx], cnt / len(column)))
        # LR
        plt.subplot(1,2,2)
        plt.hist(lr_col)
        plt.axvline(x = lr_trg, linewidth=4, color='r')
        plt.title("LR {0} coeff pvalue {1:.4f}".format(columnName[idx], lr_cnt / len(lr_col)))
        plt.savefig('PT-{0}.png'.format(idx), format='png')
    
    
    
    
    


def crimeRegression_eachCategory(year=2010):
    header = ['ARSON', 'ASSAULT', 'BATTERY', 'BURGLARY', 'CRIM SEXUAL ASSAULT', 
    'CRIMINAL DAMAGE', 'CRIMINAL TRESPASS', 'DECEPTIVE PRACTICE', 
    'GAMBLING', 'HOMICIDE', 'INTERFERENCE WITH PUBLIC OFFICER', 
    'INTIMIDATION', 'KIDNAPPING', 'LIQUOR LAW VIOLATION', 'MOTOR VEHICLE THEFT', 
    'NARCOTICS', 'OBSCENITY', 'OFFENSE INVOLVING CHILDREN', 'OTHER NARCOTIC VIOLATION',
    'OTHER OFFENSE', 'PROSTITUTION', 'PUBLIC INDECENCY', 'PUBLIC PEACE VIOLATION',
    'ROBBERY', 'SEX OFFENSE', 'STALKING', 'THEFT', 'WEAPONS VIOLATION', 'total']
    W = generate_transition_SocialLag(year)
    i = retrieve_income_features()
    e = retrieve_education_features()
    r = retrieve_race_features()
    predCrimes = {}
    unpredCrimes = {}
    for idx, val in enumerate(header):
        Y = retrieve_crime_count(year, idx+1)
        
        f1 = np.dot(W, Y)
        f = np.concatenate( (f1, np.ones(f1.shape)), axis=1 )
        Y = Y.reshape((77,))
        
        # linearRegression(f1, Y)
        cnt = 0
        for z in Y:
            if z == 0:
                cnt += 1
        print ",".join( [val, str(cnt), ""] ),  # sparseness
        res = negativeBinomialRegression(f, Y)
        if res.mle_retvals['converged']:
            predCrimes[val] = [cnt, len(Y)]
        else:
            unpredCrimes[val] = [cnt, len(Y)]
        
    return predCrimes, unpredCrimes
    
    
    
    
    
if __name__ == '__main__':
    # generate_geographical_SocialLag('../data/chicago-CA-geo-neighbor')
   
#   crimeRegression_eachCategory()
    # f = unitTest_onChicagoCrimeData()
#   print f.summary()

#    header = ['ARSON', 'ASSAULT', 'BATTERY', 'BURGLARY', 'CRIM SEXUAL ASSAULT', 
#    'CRIMINAL DAMAGE', 'CRIMINAL TRESPASS', 'DECEPTIVE PRACTICE', 
#    'GAMBLING', 'HOMICIDE', 'INTERFERENCE WITH PUBLIC OFFICER', 
#    'INTIMIDATION', 'KIDNAPPING', 'LIQUOR LAW VIOLATION', 'MOTOR VEHICLE THEFT', 
#    'NARCOTICS', 'OBSCENITY', 'OFFENSE INVOLVING CHILDREN', 'OTHER NARCOTIC VIOLATION',
#    'OTHER OFFENSE', 'PROSTITUTION', 'PUBLIC INDECENCY', 'PUBLIC PEACE VIOLATION',
#    'ROBBERY', 'SEX OFFENSE', 'STALKING', 'THEFT', 'WEAPONS VIOLATION', 'total']
#    
#    errors = np.zeros((9, len(header)))
#    mre1 = np.zeros((9, len(header)))
#    mre2 = np.zeros((9, len(header)))
#    for idx, val in enumerate(header):
#        for j in range(9):
#            r1 = leaveOneOut_evaluation_onChicagoCrimeData(2010, ['corina'], crime_idx=idx+1, flow_type=j)
#            r2 = leaveOneOut_evaluation_onChicagoCrimeData(2010, ['corina', 'sociallag'], crime_idx=idx+1, flow_type=j)
#            mre1[j][idx] = r1[0,2]
#            mre2[j][idx] = r2[0,2]
#            errors[j][idx] = r1[0,2] - r2[0,2]
#    np.savetxt('errors.array', errors)
#    np.savetxt('mre1.array', mre1)
#    np.savetxt('mre2.array', mre2)
    
    
    leaveOneOut_evaluation_onChicagoCrimeData(2010, ['corina'], verboseoutput=True)
#    permutationTest_onChicagoCrimeData(2010, ['corina', 'sociallag'])